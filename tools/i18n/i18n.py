import os
import re
import sys
import argparse
import ConfigParser
from collections import OrderedDict
from xlsxwriter import Workbook
from xlsxwriter.utility import xl_rowcol_to_cell
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# TODO: handle an array with object-like items
# TODO: this array parsing code is a god damned mess. either clean this up or use PLY or some other tokenizer.
def parse_array(s, i):
    if s[i] != '(':
        raise RuntimeError('invalid array syntax')
    i += 1
    items = []
    while True:
        r = parse_array_item(s, i)
        if r is None:
            print 'Array parse failure for in string "{}"'.format(s)
            break
        i, item = r
        items.append(item)
        if s[i] == ',':
            i += 1
            while s[i] == ',':
                i += 1
                items.append('')
            if s[i] == ')':
                # end of list
                break
            continue
        for j, c in enumerate(s[i:]):
            # skip whitespace
            if c != ' ':
                i += j
                break
        if s[i] == ')':
            # end of the list
            break
        i += 1
    return items


def parse_array_item(s, i):
    if s[i] == '\"':
        i += 1
        st = ''
        j = 0
        for j, c in enumerate(s[i:]):
            if c == '\"':
                # string ended, break out
                break
            else:
                st += c
        return i + j + 1, st
    elif s[i] == '(':
        # TODO: parse structure
        # return len(s) - 1, s[i + 1:]
        return None
    else:
        return None

def read_localization_file(sections, language, path):
    config = ConfigParser.ConfigParser()
    config.optionxform = str
    config.read(path)
    # iterate over sections
    for section_name in config.sections():
        # add section if it doesn't already exist
        if section_name not in sections:
            sections[section_name] = OrderedDict()
        section = sections[section_name]
        # add key if it doesn't already exist
        for item in config.items(section_name):
            key, value = item
            if key not in section:
                section[key] = dict()
            # update key language value
            match = re.match(r'^\"(.+)\"$', value)
            if match is not None:
                value = [match.group(1)]
            else:
                match = re.match(r'^\((.+)\)$', value)
                # TODO: split these
                if match is None:
                    pass
                else:
                    value = parse_array(value, 0)
            if value is not None:
                values = section[key]
                values[language] = value
    return sections


def main():

    # parse options
    argparser = argparse.ArgumentParser()
    argparser.add_argument('dir', default='.', help='root directory')
    argparser.add_argument('-mod', required=True, help='mod name')
    argparser.add_argument('-digest', action='store_true', required=False, default=False)  # TODO: turn into options
    argparser.add_argument('-output', action='store_true', required=False, default=False)
    args = argparser.parse_args()

    '''
    # do a clean recompile of the project and run dumpint
    if subprocess.call(['python', os.path.join(args.dir, 'tools', 'make', 'make.py'), '-dumpint', '-clean', '-mod', 'DarkestHourDev', args.dir]) != 0:
        raise RuntimeError
    '''

    args.dir = os.path.abspath(args.dir)

    if not os.path.isdir(args.dir):
        print 'error: "{}" is not a directory'.format(args.dir)
        sys.exit(1)

    # system directory
    sys_dir = os.path.join(args.dir, 'System')

    if not os.path.isdir(sys_dir):
        print 'error: could not resolve System directory'
        sys.exit(1)

    # mod directory
    mod_dir = os.path.join(args.dir, args.mod)

    if not os.path.isdir(mod_dir):
        print 'error: could not resolve mod directory'
        sys.exit(1)

    # mod system directory
    mod_sys_dir = os.path.join(mod_dir, 'System')

    if not os.path.isdir(mod_sys_dir):
        print 'error could not resolve mod system directory'
        sys.exit(1)

    workbook_name = '{0}.xlsx'.format(args.mod)

    if args.output:
        # TODO: write to mod folder
        workbook = Workbook(workbook_name)

        # TODO: set up formats
        section_format = workbook.add_format({'bg_color': '#808080', 'font_color': '#ffffff'})
        value_format = workbook.add_format({'text_wrap': True})
        languages_format = workbook.add_format({'locked': True, 'bg_color': '#000000', 'font_color': '#ffffff', 'bold': True})
        key_format = workbook.add_format({'locked': True, 'bold': True, 'bg_color': '#c0c0c0'})
        int_value_format = workbook.add_format({'bg_color': '#ffffe0', 'text_wrap': True})
        no_data_format = workbook.add_format({'bg_color': '#ffcccb', 'text_wrap': True})
        data_format = workbook.add_format({'bg_color': '#ccffcb', 'text_wrap': True})

        # read i18n configuration file
        with open(os.path.join(mod_dir, '.i18n')) as f:
            languages = f.read().splitlines()

        default_language = languages[0]

        # gather default language files
        packages = filter(lambda x: x.endswith('.{0}'.format(default_language)), os.listdir(mod_sys_dir))
        for package in packages:
            # read localization files
            sections = OrderedDict()
            for language in languages:
                path = os.path.join(mod_sys_dir, os.path.splitext(package)[0] + '.' + language)
                sections = read_localization_file(sections, language, path)

            # strip sections files down
            sections_to_remove = []
            for section in sections.keys():
                # go through keys
                keys_to_delete = []
                for key in sections[section].keys():
                    values = sections[section][key]
                    if default_language not in values.keys():
                        keys_to_delete.append(key)
                for key in keys_to_delete:
                    del sections[section][key]
                if len(sections[section]) == 0:
                    sections_to_remove.append(section)
            for section in sections_to_remove:
                del sections[section]

            worksheet = workbook.add_worksheet(os.path.splitext(package)[0])
            worksheet.freeze_panes(1, 0)
            worksheet.freeze_panes(0, 2)
            worksheet.set_column('A:A', 32)
            worksheet.set_column('B:Z', 64, value_format)
            col = 1
            # TODO: go through all the languages, write them starting at column 2
            for lang in languages:
                worksheet.write(0, col, lang, languages_format)
                col += 1
            row = 1
            for section in sections.keys():
                col = 0
                worksheet.set_row(row, None, section_format)
                worksheet.write(row, col, section)
                row += 1
                for key in sections[section].keys():
                    col = 0
                    # TODO: split into multiple records depending on formatting
                    if default_language in sections[section][key].keys():
                        # write the key name
                        worksheet.write(row, col, key, key_format)

                        # TODO; go through through the values of the BASE one, iterate row, do look-ups (left-to-right) for other languages
                        for i, value in enumerate(sections[section][key][default_language]):
                            col = 1
                            worksheet.write(row, col, value.decode('cp1252'), int_value_format)
                            for language in languages[1:]:
                                col += 1
                                if language not in sections[section][key]:
                                    # TODO: no value, mark red
                                    pass
                                else:
                                    vals = sections[section][key][language]
                                    if i >= 0 and i < len(vals):
                                        worksheet.write(row, col, vals[i].decode('cp1252'))
                                cell = xl_rowcol_to_cell(row, col)
                                worksheet.conditional_format(cell, {'type': 'blanks', 'format': no_data_format})
                                worksheet.conditional_format(cell, {'type': 'no_blanks', 'format': data_format})
                            row += 1
        workbook.close()

    if args.digest:
        # now READ IT IN
        print 'Loading workboook ' + workbook_name + '...'
        workbook = load_workbook(workbook_name)
        print 'Workbook loaded!'
        for sheet_name in workbook.get_sheet_names():
            sheet = workbook.get_sheet_by_name(sheet_name)
            # gather up languages in the file
            col = 2
            row = 1
            languages = []
            while True:
                cell = '{0}{1}'.format(get_column_letter(col), row)
                if sheet[cell].value is None:
                    break
                languages.append(sheet[cell].value)
                col += 1
            # TODO: go down the rows, find the first section:
            # sections are ones where the first column is filled, but there's nothing in the `int` column
            row += 1
            col = 1
            sections = OrderedDict()
            section = None
            key = None
            while True:
                cell = 'A{0}'.format(row)
                cell2 = 'B{0}'.format(row)
                if sheet[cell].value is None and sheet[cell2].value is None:
                    # EOF, basically
                    break
                if sheet[cell].value is not None and sheet[cell2].value is None:
                    # new section, add it!
                    section = OrderedDict()
                    sections[sheet[cell].value] = section
                else:
                    # we got a key/value row on our hands here
                    if sheet[cell].value is not None:
                        key = sheet[cell].value
                    col = 2
                    for language in languages:
                        # go searching for the translations across the rows, do a look-up for the cell
                        cell = '{0}{1}'.format(get_column_letter(col), row)
                        if key not in section:
                            # add key to section
                            section[key] = dict()
                        if language not in section[key]:
                            section[key][language] = list()

                        value = sheet[cell].value
                        section[key][language].append(value)
                        col += 1
                row += 1
            # now write out translations to their respective files
            for language in languages[1:]:
                file_path = os.path.join(mod_sys_dir, sheet_name + '.' + language)
                with open(file_path, 'wb') as f:
                    # iterate over each section
                    for section_name in sections.keys():
                        # we keep a list of key-value pairs to write out, we can query
                        # the length of the list later on and decide whether to write
                        # the section out at all
                        key_values = []
                        for key in sections[section_name].keys():
                            # make sure present language has anything for this key
                            if language in sections[section_name][key]:
                                # fetch the translated values
                                values = sections[section_name][key][language]
                                # fetch the original values
                                original_values = sections[section_name][key]['int']    # TODO: remove magic string
                                # check length of values, parse depending on size
                                if len(values) == 1 and values[0] is not None:
                                    key_values.append((key, '"' + match_whitespace(original_values[0], values[0]).encode('cp1252') + '"'))
                                else:
                                    # trasform the values to string representations; `None`s become empty strings
                                    strings = map(lambda x: '"' + x.encode('cp1252') + '"' if x is not None else '', values)
                                    # trim rightside of empty strings, as it is redundant to write out joining commas for empty values
                                    j = -1
                                    for i in xrange(len(strings) - 1, -1, -1):
                                        if strings[i] != '':
                                            j = i
                                            break
                                    strings = strings[0:j + 1]
                                    if len(strings) > 0:
                                        key_values.append((key, '(' + ','.join(strings) + ')'))
                        if len(key_values) == 0:
                            continue
                        f.write('[{0}]\n'.format(section_name))
                        for key, value in key_values:
                            f.write('{0}={1}\n'.format(key, value))
                        f.write('\n')
                if os.stat(file_path).st_size == 0:
                    os.remove(file_path)




def match_whitespace(original, translated):
    trailing = len(original) - len(original.lstrip(' '))
    if len(original.strip(' ')) > 0:
        leading = len(original) - len(original.rstrip(' '))
    return (' ' * trailing) + translated.strip(' ') + (' ' * leading)

if __name__ == "__main__":
    main()
